import time
import re
from urllib.request import urlopen
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import ElementNotVisibleException
from bs4 import BeautifulSoup
import os
import sys
import errno
import glob
import os.path
import datetime
from http.client import RemoteDisconnected
import platform
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from db.issue import IssueHandler
from utilities.helper import Helper


class Fetcher:

    __pdf_source = ""
    __paorg_source = ""
    # Holds the id's of the issues that may appear on the search page
    __possible_issues = range(1, 16)
    __driver = None

    download_folder = 'pdfs'

    def __init__(self, pdf_source):
        self.__pdf_source = pdf_source
        self.__paorg_source = 'http://www.minfin.gr/epopteuomenoi-phoreis'
        self.download_links = []
        self.download_folder = os.path.join(os.getcwd(), self.download_folder)

        chromeOptions = self.set_download_def_dir_prefs()
        self.set_driver(chromeOptions)
        
        self.__issue_handler = IssueHandler()

    def set_driver(self, chromeOptions):
        if platform.system() != 'Windows':
            self.__driver = webdriver.Chrome(os.path.join(os.path.join(os.path.dirname(__file__), ".."), "../drivers/chromedriver"),
                                            options=chromeOptions)
        else:
            self.__driver = webdriver.Chrome(os.path.join(os.path.join(os.path.dirname(__file__), ".."), "../drivers/chromedriver.exe"),
                                            options=chromeOptions)

    def set_download_folder(self, download_dir):
        self.download_folder = os.path.join(os.getcwd(), download_dir)

    def set_download_def_dir_prefs(self):
        chromeOptions = webdriver.ChromeOptions()
        prefs = {"download.default_directory": os.getcwd() + "\\" + self.download_folder}
        chromeOptions.add_experimental_option("prefs", prefs)
        return chromeOptions

    def reset_download_settings(self, download_dir):
        self.set_download_folder(download_dir)
        self.set_driver(self.set_download_def_dir_prefs())

    def file_exists(self, directory, file_name, file_extension = 'pdf'):
        return os.path.isfile(os.getcwd() + "\\" + directory + "\\" + file_name + '.' + file_extension)

    def find_start_number(self, type, year):
        conditions = {"type": [type], "date": [Helper.date_to_unix_timestamp(year), '>='],
                      "issues.date": [Helper.date_to_unix_timestamp(str(int(year) + 1)), '<']}
        issues = self.__issue_handler.load_all(conditions=conditions)

        return len(issues)

    def download_all_issues(self, type, year):

        driver = self.__driver
        driver.get(self.__pdf_source)
        print("@@@@@@@@@@@@@@@")
        # Indicates whether or not more searches will be needed to find all results
        additional_issues = True

        # Find the issue type
        issue_type = driver.find_element_by_id("label-issue-id-" + str(type)).text

        # Indicates at which issue the next search must start
        num_start = self.find_start_number(issue_type, str(year))

        # Selects the year of the issues
        year_select = Select(driver.find_element_by_name("year"))
        year_select.select_by_value(str(year))

        try:
            # Checks the box for this certain type of issues
            driver.find_element_by_name("chbIssue_" + str(type)).click()
        except ElementNotVisibleException as e:
            print("This type of issues is not available.")
            return

        while additional_issues:

            # Filters the search by issue number
            driver.find_element_by_name("fekNumberFrom").clear()
            driver.find_element_by_name("fekNumberFrom").send_keys(str(num_start))
            driver.find_element_by_name("fekNumberTo").clear()
            driver.find_element_by_name("fekNumberTo").send_keys(str(num_start + 200))

            # Submits the search form
            driver.find_element_by_name("search").click()

            # Waits 1 second for the results to load
            time.sleep(1)

            count = 0
            num_results = 0

            # Gets the amount of results from the messages displayed.
            for element in driver.find_elements_by_xpath('//div[@class="non-printable"]'):
                count += 1
                if "Βρέθηκαν" in element.text:
                    # @todo: Try regex instead of search loop and do some benchmarking
                    words = element.text.split(" ")
                    for index, word in enumerate(words):
                        if word == "αποτελέσματα":
                            num_results = int(words[index-1])
                    break

            # If there are no results for the search we abort
            if num_results == 0:
                print("No results have been found")
                break

            # The maximum number of results is 200, so if the result contains 200 an additional search will be needed
            if num_results < 200:
                additional_issues = False
            else:
                num_start += 200

            # By default we'll see the first page of results, well.. first
            active_page = 1

            # Gets the pagination list items
            pages = driver.find_elements_by_class_name("pagination_field")
            # If there's no paginations then there's one page (max)
            num_pages = len(pages) if len(pages) else 1

            for current_page in range(0, num_pages):

                # Extract and handle download links.
                self.extract_download_links(driver.page_source, issue_type)

                # We have to re-find the pagination list because the DOM has been rebuilt.
                pages = driver.find_elements_by_class_name("pagination_field")
                # Loads the next page of results
                if current_page + 1 < len(pages):
                    pages[current_page + 1].click()
                    time.sleep(1)

    def handle_download(self, download_page, params):

        try:
            # First we get the redirect link from the download page
            html = Helper.get_url_contents(download_page)
            beautiful_soup = BeautifulSoup(html, "html.parser")
            meta = beautiful_soup.find("meta", {"http-equiv": "REFRESH"})
            download_link = meta['content'].replace("0;url=", "")

            # We do the same process twice because it involves 2 redirects.
            beautiful_soup = BeautifulSoup(Helper.get_url_contents(download_link), "html.parser")
            meta = beautiful_soup.find("meta", {"http-equiv": "REFRESH"})
            download_link = meta['content'].replace("0;url=", "")
        except RemoteDisconnected as e:
            print(e)
            self.__issue_handler.create(params['issue_title'], params['issue_type'], params['issue_number'],
                                        'N/A', params['issue_date'])
            return

        if Helper.download(download_link, params['issue_title'] + ".pdf", self.download_folder):
            issue_file = os.path.join(self.download_folder, params['issue_title'] + ".pdf")
            self.__issue_handler.create(params['issue_title'], params['issue_type'], params['issue_number'],
                                        issue_file, params['issue_date'])

    def extract_download_links(self, html, issue_type):
        beautiful_soup = BeautifulSoup(html, "html.parser")
        result_table = beautiful_soup.find("table", {"id": "result_table"})
        rows = result_table.find_all("tr")

        if result_table.find("ul", {"id": "sitenav"}):
            start_row = 2
            end_row = -1
        else:
            start_row = 1
            end_row = len(rows)

        # We ignore the first 2 rows if there's pagination or the first row if there's not and the last one
        for row in rows[start_row:end_row]:
            cells = row.find_all("td")
            info_cell = cells[1].find("b")
            download_cell = cells[2].find_all("a")

            info_cell_text = info_cell.get_text()
            info_cell_text = ' '.join(info_cell_text.split())
            info_cell_parts = info_cell_text.split(" - ")

            issue_title = info_cell_text
            issue_date = info_cell_parts[1]
            issue_title_first = issue_title.split("-")[0]
            issue_number = re.search(pattern=r'\d+', string=issue_title_first).group(0)

            # Skip saved items
            if self.__issue_handler.load_by_title(issue_title):
                continue

            date_parts = issue_date.split(".")
            issue_unix_date = datetime.datetime(day=int(date_parts[0]), month=int(date_parts[1]),
                                                year=int(date_parts[2]))

            download_path = download_cell[1]['href'] if len(download_cell) > 1 else download_cell[0]['href']
            download_link = "http://www.et.gr" + download_path
            params = {"issue_title": issue_title, "issue_date": issue_unix_date, "issue_number": issue_number,
                      "issue_type": issue_type}
            self.handle_download(download_link, params)

    def scrape_pdfs(self, year_start=2016, year_end=2017):

        # Set custom download_folder & reset pre-existing settings 
        if year_end > year_start:
            download_dir = str(year_start) + '_to_' + str(year_end) + '_' + 'issues'
            download_dir = '../data/' + download_dir
        else:
            download_dir = '../data/' + str(year_start) + '_' + 'issues'

        self.reset_download_settings(download_dir)

        # Creates the folder if not exists
        try:
            os.makedirs(download_dir)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

        # Download issue pdfs
        for year in range(year_start, year_end + 1):
            for i in self.__possible_issues:
                self.download_all_issues(i, year)

    def scrape_paorgs(self, local_files):
        # From local data 
        files_location = '../data/NE_resources' 
        local_PAOrgs = []
        for file in local_files:
            file_path = files_location + '/' + file
            wb = load_workbook(filename=file_path)
            ws = wb.active
            # PAOrgs in col C for currently tested xlsx docs
            PAOrg_col = ws['C']
        
            for cell in PAOrg_col:
                paorg = cell.value
                # Strip paorg of possible acronyms & append them both to list
                parentheses = re.findall(r'\([^)]*\)', paorg)
                for parenth_elem in parentheses:
                    paorg = paorg.replace(parenth_elem, '')
                    parenth_elem = parenth_elem.replace('(', '').replace(')', '')
                    local_PAOrgs.append(parenth_elem)
                local_PAOrgs.append(paorg)

            local_PAOrgs = list(set(local_PAOrgs))

        # From web data
        src_html = urlopen(self.__paorg_source)
        soup = BeautifulSoup(src_html, 'html.parser')

        web_PAOrgs = []
        ep_foreis_table = soup.find('table', attrs={'border':"0", 'cellpadding':"0", 'cellspacing':"0", 'width':"582"})

        ep_foreis_body = ep_foreis_table.find('tbody')
        ep_foreis = ep_foreis_body.find_all('tr')

        for row in ep_foreis:
            cols = row.find_all('td')
            cols = [elem.text.strip() for elem in cols]
            
            # Get paorg (non-empty values)
            paorg = [elem for elem in cols if elem]
            paorg = ''.join(paorg)
            
            # Strip paorg of possible acronyms & append them both to list
            parentheses = re.findall(r'\([^)]*\)', paorg)
            for parenth_elem in parentheses:
                paorg = paorg.replace(parenth_elem, '')
                parenth_elem = parenth_elem.replace('(', '').replace(')', '')
                web_PAOrgs.append(parenth_elem)
            web_PAOrgs.append(paorg)

        return list(set(local_PAOrgs + web_PAOrgs))

    # Fetch public administration organizations
    def fetch_paorgs(self, local_files):
        try:
            PAOrgs = self.scrape_paorgs(local_files)
        except FileNotFoundError:
            raise

        return PAOrgs

    # Fetch responsibility assignments
    def fetch_respas():
        assignment_verbs = []
        pass 